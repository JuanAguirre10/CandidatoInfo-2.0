from openpyxl import load_workbook
from django.db import transaction
import logging
import re

logger = logging.getLogger(__name__)


def import_from_excel(file, model, field_mapping):
    try:
        wb = load_workbook(file, data_only=True, read_only=False)
        ws = wb.active
        
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        if not headers:
            return {
                'created': 0,
                'updated': 0,
                'errors': ["El archivo no tiene encabezados válidos"],
                'total_processed': 0
            }
        
        logger.info(f"Headers encontrados: {headers}")
        
        created_count = 0
        updated_count = 0
        errors = []
        
        integer_fields = ['poblacion', 'electores_registrados', 'numero_diputados', 'numero_senadores', 'fundacion_año', 'numero_lista', 'edad', 'posicion_lista', 'numero_preferencial']
        coordinate_fields = ['latitud', 'longitud']
        foreign_key_fields = ['partido_id', 'circunscripcion_id']
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            
            try:
                data = {}
                
                for idx, value in enumerate(row):
                    if idx >= len(headers):
                        break
                    
                    header = headers[idx]
                    field_name = field_mapping.get(header)
                    
                    if field_name:
                        if value == '' or value is None or (isinstance(value, str) and value.strip() == ''):
                            data[field_name] = None
                        else:
                            if isinstance(value, str):
                                value = value.strip()
                            
                            
                            
                            if field_name == 'estado':
                                value_lower = str(value).lower().strip()
                                if value_lower in ['completado', 'completed', 'finalizado']:
                                    data[field_name] = 'completado'
                                elif value_lower in ['en_ejecucion', 'en ejecucion', 'en ejecución', 'en curso', 'ejecutando', 'in progress']:
                                    data[field_name] = 'en_ejecucion'
                                elif value_lower in ['suspendido', 'suspended', 'pausado']:
                                    data[field_name] = 'suspendido'
                                elif value_lower in ['activo', 'active']:
                                    data[field_name] = 'activo'
                                elif value_lower in ['inactivo', 'inactive']:
                                    data[field_name] = 'inactivo'
                                elif value_lower in ['inhabilitado', 'disabled']:
                                    data[field_name] = 'inhabilitado'
                                elif value_lower in ['inscrito', 'registered']:
                                    data[field_name] = 'inscrito'
                                elif value_lower in ['observado', 'observed']:
                                    data[field_name] = 'observado'
                                elif value_lower in ['excluido', 'excluded']:
                                    data[field_name] = 'excluido'
                                elif value_lower in ['aprobado', 'approved']:
                                    data[field_name] = 'aprobado'
                                else:
                                    data[field_name] = 'completado'
                            
                            elif field_name == 'color_principal':
                                color_value = str(value).strip()
                                if re.match(r'^#[0-9A-Fa-f]{6}$', color_value):
                                    data[field_name] = color_value
                                elif len(color_value) <= 50:
                                    data[field_name] = color_value
                                else:
                                    data[field_name] = '#3b82f6'
                            
                            elif field_name == 'partido_id':
                                try:
                                    partido_value = str(value).strip()
                                    if partido_value.isdigit():
                                        data[field_name] = int(partido_value)
                                    else:
                                        from partidos.models import PartidoPolitico
                                        partido = PartidoPolitico.objects.filter(siglas__iexact=partido_value).first()
                                        data[field_name] = partido.id if partido else None
                                except Exception:
                                    data[field_name] = None
                            
                            elif field_name == 'circunscripcion_id':
                                try:
                                    circ_value = str(value).strip()
                                    if circ_value.isdigit():
                                        data[field_name] = int(circ_value)
                                    else:
                                        from circunscripciones.models import Circunscripcion
                                        circ = Circunscripcion.objects.filter(codigo__iexact=circ_value).first()
                                        data[field_name] = circ.id if circ else None
                                except Exception:
                                    data[field_name] = None
                            
                            elif field_name in integer_fields:
                                try:
                                    clean_value = str(value).replace(' ', '').replace('\u202f', '').replace(',', '')
                                    data[field_name] = int(clean_value) if clean_value else None
                                except (ValueError, AttributeError):
                                    data[field_name] = None
                            
                            elif field_name in coordinate_fields:
                                try:
                                    data[field_name] = float(value) if value else None
                                except (ValueError, TypeError):
                                    data[field_name] = None
                            
                            else:
                                data[field_name] = value
                
                data_sin_id = {k: v for k, v in data.items() if k != 'id' and v is not None}
                
                if not data_sin_id:
                    logger.warning(f"Fila {row_num} saltada - sin datos válidos")
                    continue
                
                with transaction.atomic():
                    obj_id = data.pop('id', None)
                    
                    if obj_id:
                        try:
                            obj = model.objects.get(id=obj_id)
                            for field, value in data.items():
                                setattr(obj, field, value)
                            obj.save()
                            updated_count += 1
                            logger.info(f"Registro {obj_id} actualizado")
                        except model.DoesNotExist:
                            obj = model.objects.create(**data)
                            created_count += 1
                            logger.info(f"Registro creado (ID {obj_id} no existía)")
                    else:
                        obj = model.objects.create(**data)
                        created_count += 1
                        logger.info(f"Nuevo registro creado: {obj.id}")
                
            except Exception as e:
                error_msg = f"Fila {row_num}: {str(e)}"
                errors.append(error_msg)
                logger.error(error_msg, exc_info=True)
                continue
        
        result = {
            'created': created_count,
            'updated': updated_count,
            'errors': errors,
            'total_processed': created_count + updated_count
        }
        
        logger.info(f"Importación completada: {result}")
        return result
        
    except Exception as e:
        error_msg = f"Error al procesar archivo: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return {
            'created': 0,
            'updated': 0,
            'errors': [error_msg],
            'total_processed': 0
        }