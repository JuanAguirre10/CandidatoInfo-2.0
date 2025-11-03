from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from datetime import datetime


def export_to_excel(data, filename, headers, title):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    
    title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(color="FFFFFF", bold=True, size=12)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = title_fill
        cell.font = title_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row_num, row_data in enumerate(data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response